from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from openpyxl import workbook

path_to_save_file = fr'C:\Users\luane\Downloads\valores.xlsx'
lista_acoes       = ['aure3', 'cxse3', 'amer3']

servico = Service(ChromeDriverManager().install())
chrome = webdriver.Chrome(service=servico)
chrome.implicitly_wait(30) #espera a pagina carregar até 30s antes de dar erro
chrome.maximize_window()

wb = workbook.Workbook()
ws = wb.active
columns_names = ['PAPEL', 'PREÇO', 'P/L', 'P/VP', 'D.Y.', 'ROE', 'DIV. LIQUIDA', 'CAGR LUCRO']

for i, v in enumerate(columns_names):
    i += 1
    ws.cell(1, i, v)

for linha, ticket in enumerate(lista_acoes):
    chrome.get('https://statusinvest.com.br/acoes/' + ticket)

    PRICE     = chrome.execute_script("return document.getElementsByClassName('info special w-100 w-md-33 w-lg-20')[0].textContent.split('\\n')[5]")
    PL        = chrome.execute_script("return document.getElementsByClassName('w-50 w-sm-33 w-md-25 w-lg-16_6 mb-2 mt-2 item ')[1].textContent.split('\\n')[7]")
    PVP       = chrome.execute_script("return document.getElementsByClassName('w-50 w-sm-33 w-md-25 w-lg-16_6 mb-2 mt-2 item ')[3].textContent.split('\\n')[7]")
    DY        = chrome.execute_script("return document.getElementsByClassName('w-50 w-sm-33 w-md-25 w-lg-16_6 mb-2 mt-2 item ')[0].textContent.split('\\n')[7]")
    ROE       = chrome.execute_script("return document.getElementsByClassName('w-50 w-sm-33 w-md-25 w-lg-50  mb-2 mt-2 item')[4].textContent.split('\\n')[7]")
    DIVLIQ    = chrome.execute_script("return document.getElementsByClassName('w-50 w-sm-33 w-md-25 w-lg-16_6 mb-2 mt-2 item')[15].textContent.split('\\n')[7]")
    CAGRLUCRO = chrome.execute_script("return document.getElementsByClassName('w-50 w-sm-33 w-md-25 w-lg-50  mb-2 mt-2 item')[9].textContent.split('\\n')[4]")

    values    = [str(ticket).upper(), PRICE, PL, PVP, DY, ROE, DIVLIQ, CAGRLUCRO]

    for i, v in enumerate(values):
        i += 1
        ws.cell(linha+2, i, v)

wb.save(path_to_save_file)

